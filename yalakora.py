import requests
from bs4 import BeautifulSoup
import openpyxl
import os

class yallakora:

    def __init__(self):
        pass

    def saving_in_excel(self,data):

        filepath = "./match_info.xlsx"

        # check if the file exists
        if os.path.isfile(filepath):
            # load the existing file and add data to it
            self.workbook = openpyxl.load_workbook(filepath)
            self.worksheet = self.workbook.active
        else:
            # create a new workbook and save it to the specified path
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active

        # get the row to append data to
        row_number = self.worksheet.max_row + 1

        # write dictionary keys to the first row of the worksheet
        col = 1
        for key in data[0].keys():
            self.worksheet.cell(row=1, column=col, value=key)
            col += 1

        # write dictionary values to the worksheet
        row = 2
        for datum in data:
            col = 1
            for value in datum.values():
                self.worksheet.cell(row=row, column=col, value=value)
                col += 1
            row += 1

        # save the workbook
        self.workbook.save("match_info.xlsx")

    def main(self):
        #https://www.yallakora.com/match-center/%D9%85%D8%B1%D9%83%D8%B2-%D8%A7%D9%84%D9%85%D8%A8%D8%A7%D8%B1%D9%8A%D8%A7%D8%AA?date=5/8/2023#days
        page  = requests.get('https://www.yallakora.com/match-center/%d9%85%d8%b1%d9%83%d8%b2-%d8%a7%d9%84%d9%85%d8%a8%d8%a7%d8%b1%d9%8a%d8%a7%d8%aa#nav-menu')

        src = page.content
        soup = BeautifulSoup(src, 'lxml')

        chap = soup.find_all('div', {'class': 'matchCard'})
        info = []
        for div in chap:
            ligue = div.find_all('li')
            title = div.find('h2').text.strip()
            for match in ligue:
                info.append({
                    'Title' : title,
                    'Team A' : match.find('div', {'class': 'teamA'}).find('p').text.strip(),
                    'Team B' : match.find('div', {'class': 'teamB'}).find('p').text.strip(),
                    'Score' : match.find_all('span', {'class': 'score'})[0].text+' - '+match.find_all('span', {'class': 'score'})[1].text,
                    'Time' : match.find('div', {'class': 'MResult'}).find('span', {'class': 'time'}).text
                    })

        self.saving_in_excel(info) 

    

yallakora().main()